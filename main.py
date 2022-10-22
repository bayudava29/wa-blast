import pywhatkit
import openpyxl
from datetime import datetime

# get excel file
workbook = openpyxl.load_workbook('resources/contact.xlsx')
sheet = workbook.active
max_row = sheet.max_row

def parse_message(name, number):
    return """
Halo %s (%s)
Ini pesan testing
""" % (name, number)

for i in range(2, max_row + 1):
    # time now
    now = datetime.now()
    hour = int(now.strftime("%H"))
    minute = int(now.strftime("%M")) + 1

    # get cell
    cell_wa_number = sheet.cell(row = i, column = 1)
    cell_name = sheet.cell(row = i, column = 2)

    # parse message
    message = parse_message(cell_name.value, cell_wa_number.value)

    # send to wa
    pywhatkit.sendwhatmsg(cell_wa_number.value, message, hour, minute)